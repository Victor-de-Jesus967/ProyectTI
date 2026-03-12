import glob
import sys
import os
from pathlib import Path
from PyPDF2 import PdfWriter, PdfReader, PageObject
from PyQt5.QtWidgets import QMessageBox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from PyQt5 import QtCore, QtGui, QtWidgets

import sqlite3
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet

from reportlab.platypus import SimpleDocTemplate, Paragraph, PageBreak

from Acciones import Acciones
from ActualizarBD import ActualizarBD

class Ui_MainWindow(object):
    def __init__(self):
        self.file_manager = FileManager()
        self.tableView = None
        # Instancia de la clase Acciones
        self.acciones = Acciones()  # Pasamos los objetos que necesiten, como el db_manager
        self.actualizarbd = ActualizarBD()  # Instancia para manejar la base de datos

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(960, 725)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")

        # Crea el layout principal
        self.mainLayout = QtWidgets.QVBoxLayout(self.centralwidget)

        # Layout para el título
        self.titleLayout = QtWidgets.QHBoxLayout()
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(23)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.titleLayout.addStretch()
        self.titleLayout.addWidget(self.label_3)
        self.titleLayout.addStretch()
        self.mainLayout.addLayout(self.titleLayout)

        # Layout para los campos de entrada
        self.inputLayout = QtWidgets.QGridLayout()
        self.label = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.inputLayout.addWidget(self.label, 0, 0)
        self.Lficha = QtWidgets.QLineEdit(self.centralwidget)
        self.Lficha.setObjectName("Lficha")
        self.inputLayout.addWidget(self.Lficha, 0, 1)
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.inputLayout.addWidget(self.label_4, 1, 0)
        self.comboBox = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItems([
            "Equipo Entregado Adquisición", "Arrendado Retirar",
            "Equipo Nuevo", "Equipos Reasignados", "Retiro Equipo"
        ])
        self.inputLayout.addWidget(self.comboBox, 1, 1)
        self.mainLayout.addLayout(self.inputLayout)

        # Layout para los botones
        self.buttonLayout = QtWidgets.QGridLayout()
        self.Bbuscar = QtWidgets.QPushButton("Buscar", self.centralwidget)
        self.Bbuscar.setObjectName("Bbuscar")
        self.buttonLayout.addWidget(self.Bbuscar, 0, 0)
        self.Breporte = QtWidgets.QPushButton("Reporte", self.centralwidget)
        self.Breporte.setObjectName("Breporte")
        self.buttonLayout.addWidget(self.Breporte, 0, 1)
        self.Bestatus = QtWidgets.QPushButton("Estatus", self.centralwidget)
        self.Bestatus.setObjectName("Bestatus")
        self.buttonLayout.addWidget(self.Bestatus, 0, 2)
        self.Bactualizar = QtWidgets.QPushButton("Actualizar", self.centralwidget)
        self.Bactualizar.setObjectName("Bactualizar")
        self.buttonLayout.addWidget(self.Bactualizar, 1, 0)
        self.Batras = QtWidgets.QPushButton("Atrás", self.centralwidget)
        self.Batras.setObjectName("Batras")
        self.buttonLayout.addWidget(self.Batras, 1, 1)
        self.Bsalir = QtWidgets.QPushButton("Salir", self.centralwidget)
        self.Bsalir.setObjectName("Bsalir")
        self.buttonLayout.addWidget(self.Bsalir, 1, 2)
        self.mainLayout.addLayout(self.buttonLayout)

        # Etiqueta para la base de datos
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.mainLayout.addWidget(self.label_2)

        # Tabla para mostrar datos
        self.Tabla = QtWidgets.QTableView(self.centralwidget)
        self.Tabla.setObjectName("Tabla")
        self.mainLayout.addWidget(self.Tabla)

        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        # Conectar botones a funciones
        self.Bbuscar.clicked.connect(self.buscar_ficha)
        self.Breporte.clicked.connect(self.generate_report)
        self.Bsalir.clicked.connect(self.cerrar_programa)

        # Cargar toda la base de datos en el QTableView al iniciar
        self.cargar_datos_en_tabla_EquipoEntregadoAdquisicion()
        self.cargar_datos_en_tabla_ArrendadoRetirar()
        self.cargar_datos_en_tabla_EquiposReasignados()
        self.cargar_datos_en_tabla_EquipoNuevo()
        # Conectar el botón Bactualizar a la función cargar_archivo_excel
        self.Bactualizar.clicked.connect(self.actualizarbd.cargar_archivo_excel)  # Conectar el botón Bvalidar a la función validar_archivo


        # Configura el comboBox para que llame a cargar_tabla_segun_seleccion cada vez que cambie su selección
        self.comboBox.currentIndexChanged.connect(self.actualizar_tabla_por_combo)

        # Conectar el botón Bestatus
        self.Bestatus.clicked.connect(self.verificar_estatus_retiro)

        # Conectar el botón Batras para mostrar todos los datos de la tabla
        self.Batras.clicked.connect(lambda: self.cargar_datos_en_tabla_EquipoEntregadoAdquisicion())

        # Conectar el botón a la función que abrirá la nueva ventana
        #self.Batras.clicked.connect(self.acciones.cargar_datos_en_tabla)  # Pasar la tabla desde la interfaz)

        MainWindow.setCentralWidget(self.centralwidget)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Inventario"))
        self.label.setText(_translate("MainWindow", "Ficha/Usuario/SerieCPU:"))
        self.label_4.setText(_translate("MainWindow", "Seleccionar tabla:"))
        self.label_2.setText(_translate("MainWindow", "Base de datos:"))
        self.label_3.setText(_translate("MainWindow", "Inventario"))

        # Método para buscar un registro específico por ficha y mostrarlo en la tabla

    def cerrar_programa(self):
        QtWidgets.QApplication.quit()

    def buscarficha1(self):
        ficha = self.Lficha.text()
        resultado = self.acciones.buscar_registro(ficha)
        if resultado:
            self.cargar_datos_en_tabla_EquipoEntregadoAdquisicion(resultado)
        else:
            print("No se encontró el número de ficha o usuario.")

    # Método para buscar un registro específico por ficha y mostrarlo en la tabla
    def buscar_ficha(self):
        ficha = self.Lficha.text()
        resultados = self.acciones.buscar_registro(ficha)

        if resultados:
            # Suponiendo que 'resultados' es una lista de tuplas (filas)
            # Tomamos el primer (y presumiblemente único) resultado
            registro = resultados[0]

            # Definir los nombres de las columnas manualmente según tu base de datos
            nombres_campos = ['FechaEntrega', 'SerieCPU', 'Hostname', 'Organismo', 'Componente', 'PEMP',
                              'IPEquipo', 'SerieReplicador', 'SerieTeclado', 'SerieMonitor', 'Ficha',
                              'NombreUsuario', 'Estatus', 'Ubicacion', 'Piso', 'AreaDepartamento',
                              'Subdireccion', 'ExtUsuario']

            # Transponer los datos para mostrar los campos verticalmente
            datos_verticales = list(zip(nombres_campos, registro))

            # Crear un modelo estándar para la tabla
            modelo = QtGui.QStandardItemModel()
            modelo.setHorizontalHeaderLabels(['Campo', 'Valor'])

            # Añadir los datos al modelo
            for row, (campo, valor) in enumerate(datos_verticales):
                item_campo = QtGui.QStandardItem(str(campo))
                item_valor = QtGui.QStandardItem(str(valor))
                modelo.setItem(row, 0, item_campo)
                modelo.setItem(row, 1, item_valor)

                # Asignar el modelo a la tabla
            self.Tabla.setModel(modelo)

            # Ajustar el tamaño de las columnas y filas
            self.Tabla.resizeColumnsToContents()
            self.Tabla.resizeRowsToContents()
        else:
            self.mostrar_mensaje("No se encontró el número de ficha o usuario.")

    # Método para agregar un registro llamando a Acciones
    def agregar_registro(self):
        datos = (
            # Aquí irían los datos capturados de la interfaz
        )
        self.acciones.agregar_registro(datos)
        self.cargar_datos_en_tabla_EquipoEntregadoAdquisicion()

    # Método para modificar un registro llamando a Acciones
    def modificar_registro(self):
        id_registro = self.Lficha.text()  # ID del registro a modificar
        nuevos_datos = (
            # Nuevos datos para el registro
        )
        self.acciones.modificar_registro(id_registro, nuevos_datos)
        self.cargar_datos_en_tabla_EquipoEntregadoAdquisicion()

    # Método para eliminar un registro llamando a Acciones
    def eliminar_registro(self):
        id_registro = self.Lficha.text()  # ID del registro a eliminar
        self.acciones.eliminar_registro(id_registro)
        self.cargar_datos_en_tabla_EquipoEntregadoAdquisicion()
        fila = self.tableView.currentRow()
        if fila >= 0:
            ruta_archivo = self.tableView.item(fila, 1).text()
            os.remove(ruta_archivo)  # Elimina el archivo del sistema
            self.tableView.removeRow(fila)  # Elimina la fila de la tabla

    # Función para cargar los datos en la tabla
    def cargar_datos_en_tabla_EquipoEntregadoAdquisicion(self, datos=None):
        self.acciones.cargar_datos_en_tabla_EquipoEntregadoAdquisicion(self.Tabla, datos)
        self.acciones = Acciones()
        #self.cargar_datos_en_tabla()
        # Ajustar el tamaño de las columnas y filas
        self.Tabla.resizeColumnsToContents()  # Ajustar columnas
        self.Tabla.resizeRowsToContents()  # Ajustar filas

    def cargar_datos_en_tabla_ArrendadoRetirar(self, datos=None):
        self.acciones.cargar_datos_en_tabla_ArrendadoRetirar(self.Tabla, datos)
        self.acciones = Acciones()
        #self.cargar_datos_en_tabla()
        # Ajustar el tamaño de las columnas y filas
        self.Tabla.resizeColumnsToContents()  # Ajustar columnas
        self.Tabla.resizeRowsToContents()  # Ajustar filas

    def cargar_datos_en_tabla_EquiposReasignados(self, datos=None):
        self.acciones.cargar_datos_en_tabla_EquiposReasignados(self.Tabla, datos)
        self.acciones = Acciones()
        #self.cargar_datos_en_tabla()
        # Ajustar el tamaño de las columnas y filas
        self.Tabla.resizeColumnsToContents()  # Ajustar columnas
        self.Tabla.resizeRowsToContents()  # Ajustar filas

    def cargar_datos_en_tabla_EquipoNuevo(self, datos=None):
        self.acciones.cargar_datos_en_tabla_EquipoNuevo(self.Tabla, datos)
        self.acciones = Acciones()
        #self.cargar_datos_en_tabla()
        # Ajustar el tamaño de las columnas y filas
        self.Tabla.resizeColumnsToContents()  # Ajustar columnas
        self.Tabla.resizeRowsToContents()  # Ajustar filas

    def abrir_interfaz_resguardo(self):
        # Cerrar la ventana actual
        QtWidgets.qApp.closeAllWindows()

        # Abrir la nueva ventana
        self.window = QtWidgets.QMainWindow()
        #self.ui = Interfaz2Window()
        self.ui.setupUi(self.window)
        self.window.show()

    def mostrar_mensaje(self, mensaje):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(mensaje)
        msg.setWindowTitle("Información")
        msg.exec_()

    def buscar_ficha1(self):
        ficha = self.Lficha.text()
        resultado = self.acciones.buscar_registro(ficha)

        if resultado:
            self.cargar_datos_en_tabla_EquipoEntregadoAdquisicion(resultado)
        else:
            self.mostrar_mensaje("No se encontró el número de ficha o usuario.")

    # Método para cargar los datos en la tabla, muestra todos los datos si 'datos' es None
    def cargar_datos_en_tabla1(self, datos=None):
        self.acciones.cargar_datos_en_tabla_EquipoEntregadoAdquisicion(self.Tabla, datos)

        # Ajustar el tamaño de las columnas y filas
        self.Tabla.resizeColumnsToContents()  # Ajustar columnas
        self.Tabla.resizeRowsToContents()  # Ajustar filas

    # Método para cargar la tabla según la selección en el comboBox
    # Método para limpiar la tabla
    def limpiar_tabla(self):
        # Limpia todas las filas de la tabla
        self.Tabla.setRowCount(0)

    def actualizar_tabla_por_combo(self):
        """Actualiza los datos de la tabla según la opción seleccionada en el comboBox."""
        # Obtener el texto seleccionado en el comboBox
        tabla_seleccionada = self.comboBox.currentText()

        # Llamar al método correspondiente para cargar los datos en la tabla
        if tabla_seleccionada == "Equipo Entregado Adquisición":
            print("Tabla lista Equipo Entregado Adquisición")
            self.acciones.cargar_datos_en_tabla_EquipoEntregadoAdquisicion(self.Tabla)
        elif tabla_seleccionada == "Arrendado Retirar":
            self.acciones.cargar_datos_en_tabla_ArrendadoRetirar(self.Tabla)
            print("Tabla lista Arrendado Retirar")
        elif tabla_seleccionada == "Equipo Nuevo":
            self.acciones.cargar_datos_en_tabla_EquipoNuevo(self.Tabla)
            print("Tabla lista Equipo Nuevo")
        elif tabla_seleccionada == "Retiro Equipo":
            self.acciones.cargar_datos_en_tabla_RetiroEquipo(self.Tabla)
            print("Tabla lista Equipo Nuevo")
        elif tabla_seleccionada == "Equipos Reasignados":
            self.acciones.cargar_datos_en_tabla_EquiposReasignados(self.Tabla)
            print("Tabla lista Equipos Reasignados")
        else:
            print("Tabla no reconocida.")

    def validar_archivo(self):
        criterio_busqueda = self.Lficha.text().strip()
        print(f"Criterio de búsqueda: '{criterio_busqueda}'")

        if not criterio_busqueda:
            QMessageBox.warning(None, "Advertencia", "Por favor, ingrese una ficha, fecha o número de serie.")
            return

        resultados = self.acciones.buscar_registro1(criterio_busqueda)

        if resultados:
            mensaje = "Archivos encontrados:\n\n"
            rutas_archivos = []  # Lista para almacenar las rutas de los archivos

            for item in resultados:
                registro = item['registro']
                fecha = item['fecha']
                ficha = item['ficha']
                numero_serie = item['numero_serie']

                nombre_archivo = registro[1]
                ruta_archivo = registro[2]
                rutas_archivos.append(ruta_archivo)  # Guardar la ruta

                mensaje += f"Nombre: {nombre_archivo}\n"
                mensaje += f"Ruta: {ruta_archivo}\n"
                mensaje += f"Fecha: {fecha}\n"
                mensaje += f"Ficha: {ficha}\n"
                mensaje += f"Número de Serie: {numero_serie}\n"
                mensaje += "-" * 40 + "\n"

                # Crear un QMessageBox personalizado con botones
            msgBox = QMessageBox()
            msgBox.setWindowTitle("Resultados de Búsqueda")
            msgBox.setText(mensaje)

            # Agregar botones personalizados
            abrir_button = msgBox.addButton("Abrir Archivo", QMessageBox.ActionRole)
            cancelar_button = msgBox.addButton("Cancelar", QMessageBox.RejectRole)

            msgBox.exec_()

            # Verificar qué botón se presionó
            if msgBox.clickedButton() == abrir_button:
                if len(rutas_archivos) == 1:
                    # Si solo hay un archivo, ábrelo directamente
                    try:
                        os.startfile(rutas_archivos[0])  # Para Windows
                        # Para Linux: subprocess.call(["xdg-open", rutas_archivos[0]])
                        # Para Mac: subprocess.call(["open", rutas_archivos[0]])
                    except Exception as e:
                        QMessageBox.warning(None, "Error", f"No se pudo abrir el archivo: {str(e)}")
                else:
                    # Si hay múltiples archivos, mostrar un diálogo de selección
                    seleccion, ok = QtWidgets.QInputDialog.getItem(
                        None,
                        "Seleccionar Archivo",
                        "Seleccione el archivo a abrir:",
                        [os.path.basename(ruta) for ruta in rutas_archivos],
                        0,
                        False
                    )
                    if ok and seleccion:
                        indice = [os.path.basename(ruta) for ruta in rutas_archivos].index(seleccion)
                        try:
                            os.startfile(rutas_archivos[indice])  # Para Windows
                            # Para Linux: subprocess.call(["xdg-open", rutas_archivos[indice]])
                            # Para Mac: subprocess.call(["open", rutas_archivos[indice]])
                        except Exception as e:
                            QMessageBox.warning(None, "Error", f"No se pudo abrir el archivo: {str(e)}")
        else:
            QMessageBox.warning(None, "Archivo No Encontrado", "No se encontró ningún archivo con ese criterio.")

    def verificar_estatus_retiro(self):
        try:
            # 1. Verificar si se seleccionó una fila
            indice_modelo = self.Tabla.currentIndex()
            if not indice_modelo.isValid():
                QtWidgets.QMessageBox.warning(
                    None,
                    "Advertencia",
                    "Por favor, seleccione un equipo de la tabla."
                )
                return

            modelo = self.Tabla.model()
            fila_actual = indice_modelo.row()

            # 2. Encontrar la columna "SerieCPU"
            serie_cpu_col = next(
                (col for col in range(modelo.columnCount())
                 if modelo.headerData(col, QtCore.Qt.Horizontal) == "SerieCPU"),
                -1
            )

            if serie_cpu_col == -1:
                QtWidgets.QMessageBox.warning(
                    None,
                    "Error",
                    "No se encontró la columna 'SerieCPU' en la tabla."
                )
                return

            # 3. Obtener el valor de SerieCPU
            serie_cpu = modelo.data(modelo.index(fila_actual, serie_cpu_col))
            if not serie_cpu:
                QtWidgets.QMessageBox.warning(
                    None,
                    "Error",
                    "No se pudo obtener el número de serie CPU de la fila seleccionada."
                )
                return

            # 4. Consultar la base de datos
            conexion = sqlite3.connect('equipo_computo.db')
            cursor = conexion.cursor()

            query_retiro = """
                SELECT EstatusRetiro, ObservacionesRetiro, RetiroFisico,  
                       EstatusWoo, EstatusTas  
                FROM RetiroEquipo  
                WHERE ROWID = ?  
            """
            cursor.execute(query_retiro, (fila_actual + 1,))
            resultado_retiro = cursor.fetchone()
            conexion.close()

            # 5. Construir el mensaje a mostrar
            mensaje = (
                f"<b>Información del Equipo:</b><br>"
                f"Serie CPU: {serie_cpu}<br><br>"
                f"<b>Estado del Retiro:</b><br>"
            )

            if resultado_retiro:
                mensaje += (
                    f"Estatus: {resultado_retiro[0] or 'No disponible'}<br>"
                    f"Observaciones: {resultado_retiro[1] or 'No disponible'}<br>"
                    f"Retiro Físico: {resultado_retiro[2] or 'No disponible'}<br>"
                    f"Estatus WOO: {resultado_retiro[3] or 'No disponible'}<br>"
                    f"Estatus TAS: {resultado_retiro[4] or 'No disponible'}"
                )
            else:
                mensaje += "No hay información de retiro registrada para este equipo."

            # 6. Mostrar el mensaje en un QMessageBox con estilo
            msg_box = QtWidgets.QMessageBox()
            msg_box.setWindowTitle("Estado del Retiro")
            msg_box.setTextFormat(QtCore.Qt.RichText)
            msg_box.setText(mensaje)
            msg_box.setIcon(QtWidgets.QMessageBox.Information)

            msg_box.setStyleSheet("""
                QMessageBox {
                    background-color: #f9f9f9;
                    border: 1px solid #cccccc;
                    border-radius: 8px;
                }
                QMessageBox QLabel {
                    color: #333333;
                    font-family: Arial;
                    font-size: 12px;
                    line-height: 1.6;
                }
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-size: 12px;
                    font-family: Arial;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)

            # Centrar el mensaje en la pantalla
            msg_box.setFixedWidth(450)
            msg_box.exec_()

        except Exception as e:
            QtWidgets.QMessageBox.critical(
                None,
                "Error",
                f"Ocurrió un error al verificar el estatus: {str(e)}"
            )

    def generate_report(self):
        # Mostrar un cuadro de diálogo para seleccionar el formato del reporte
        format_dialog = QtWidgets.QMessageBox()
        format_dialog.setWindowTitle("Seleccionar formato")
        format_dialog.setText("¿En qué formato deseas generar el reporte?")
        pdf_button = format_dialog.addButton("PDF", QtWidgets.QMessageBox.AcceptRole)
        excel_button = format_dialog.addButton("Excel", QtWidgets.QMessageBox.AcceptRole)
        cancel_button = format_dialog.addButton("Cancelar", QtWidgets.QMessageBox.RejectRole)

        format_dialog.exec_()

        if format_dialog.clickedButton() == pdf_button:
            self.generate_pdf_report()
        elif format_dialog.clickedButton() == excel_button:
            self.generate_excel_report()

    def generate_excel_report(self):
        # Mostrar un diálogo para seleccionar las preguntas
        dialog = QuestionSelectionDialog()
        if dialog.exec_():
            selected_questions = dialog.get_selected_questions()
            self.create_excel(selected_questions)

    def generate_pdf_report(self):
        # Mostrar un diálogo para seleccionar las preguntas
        dialog = QuestionSelectionDialog()
        if dialog.exec_():
            selected_questions = dialog.get_selected_questions()
            # Realizar las consultas y generar el reporte
            self.create_pdf(selected_questions)

    def resource_path(self, relative_path, file_manager=None):

        """Devuelve la ruta absoluta del recurso, funcionando en el entorno de desarrollo y en el ejecutable."""
        file_manager = file_manager or self.file_manager

        if hasattr(sys, '_MEIPASS'):
            # Si se ejecuta como un ejecutable generado por pyinstaller
            return os.path.join(sys._MEIPASS, relative_path)
        else:
            # Si se ejecuta en el entorno de desarrollo
            return os.path.join(os.path.abspath("."), relative_path)

    def create_pdf(self, selected_questions):
        try:

            # Verificar y limpiar archivos PDF antiguos
            self.file_manager.check_and_clean('pdf')


            # Conectar a la base de datos
            conn = sqlite3.connect('equipo_computo.db')
            cursor = conn.cursor()

            # Crear un nombre único para el PDF
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_pdf_file = f'Reporte_{timestamp}.pdf'

            # Leer el PDF de fondo
            background_pdf = self.resource_path('Encabezado.pdf')
            background = PdfReader(background_pdf)

            # Crear un documento PDF para el contenido
            temp_pdf_file = f'temp_report_{timestamp}.pdf'
            doc = SimpleDocTemplate(temp_pdf_file, pagesize=letter)
            styles = getSampleStyleSheet()
            story = [Paragraph("Reporte de Equipos de Cómputo", styles['Title']),
                     Paragraph("Fecha: " + datetime.now().strftime('%d/%m/%Y'), styles['Normal']),
                     Paragraph(" ", styles['Normal'])]

            # Orden de las preguntas
            questions_order = [
                '¿Cuantos equipos de computo se han retirado?',
                '¿Cuantos equipos de computo se han retirado del activo?',
                '¿Cuantos equipos de computo se han reasignado?',
                '¿Cuantos equipos de computo tienen replicador?',
                '¿Cuantos equipos de computo estan entregados?',
                '¿Cuantos equipos de computo estan en proceso?',
                '¿Cuantos equipos de computo hay asignados en cada contrato?',
                '¿Cuantos son laptop normal, laptop uso normal, laptop uso rudo, pc uso general, videoproyector fijo, videoproyector portatil y workstation?',
                '¿Cuantos equipos de computo hay en cada organismo?',
                '¿Cuantos equipos de computo hay en cada area?',
                '¿Cuantos equipos de computo hay en cada subdireccion del activo?'
            ]

            # Contrato y sus contabilizaciones
            Contratos = ['44', '48', '165', '165/48', '48/165']

            # Variable para controlar el número de párrafos por página
            MAX_PARAGRAPHS_PER_PAGE = 15  # Establece un límite razonable de párrafos por página
            paragraph_count = 0  # Contador de párrafos en la página actual

            # Iterar sobre preguntas en el orden especificado
            for question in questions_order:
                if question in selected_questions:  # Solo procesar preguntas seleccionadas
                    # Contar equipos retirados
                    if question == '¿Cuantos equipos de computo se han retirado?':
                        cursor.execute('''  
                                   SELECT COUNT(*)  
                        FROM RetiroEquipo  
                        WHERE EstatusRetiro IS NOT NULL  
                        AND LENGTH(TRIM(EstatusRetiro)) > 0   
                               ''')
                        count = cursor.fetchone()[0]
                        story.append(Paragraph(f'<b>{question}</b>: {count}', styles['Normal']))
                        paragraph_count += 1

                        # Contar equipos retirados del activo
                    elif question == '¿Cuantos equipos de computo se han retirado del activo?':
                        cursor.execute('''  
                                   SELECT COUNT(*)  
                                   FROM RetiroEquipo  
                                   WHERE UPPER(TRIM(RetiroFisico)) = 'RETIRADO'  
                               ''')
                        count = cursor.fetchone()[0]
                        story.append(Paragraph(f'<b>{question}</b>: {count}', styles['Normal']))
                        paragraph_count += 1

                        # Contar equipos reasignados
                    elif question == '¿Cuantos equipos de computo se han reasignado?':
                        cursor.execute('''  
                                   SELECT COUNT(*)  
                                   FROM EquiposReasignados  
                                   WHERE UPPER(TRIM(Si_No)) = 'SI'  
                               ''')
                        count = cursor.fetchone()[0]
                        story.append(Paragraph(f'<b>{question}</b>: {count}', styles['Normal']))
                        paragraph_count += 1

                        # Contar equipos con replicador
                    elif question == '¿Cuantos equipos de computo tienen replicador?':
                        cursor.execute('''  
                                   SELECT COUNT(DISTINCT SerieReplicador)  
                                   FROM EquipoEntregadoAdquisicion  
                                   WHERE SerieReplicador IS NOT NULL  
                                     AND TRIM(SerieReplicador) != ''  
                               ''')
                        count = cursor.fetchone()[0]
                        story.append(Paragraph(f'<b>{question}</b>: {count}', styles['Normal']))
                        paragraph_count += 1

                        # Contar equipos entregados
                    elif question == '¿Cuantos equipos de computo estan entregados?':
                        cursor.execute('''  
                                   SELECT COUNT(*)  
                                   FROM EquipoEntregadoAdquisicion  
                                   WHERE Estatus = 'ENTREGADO'  
                               ''')
                        count = cursor.fetchone()[0]
                        story.append(Paragraph(f'<b>{question}</b>: {count}', styles['Normal']))
                        paragraph_count += 1

                        # Contar equipos en proceso
                    elif question == '¿Cuantos equipos de computo estan en proceso?':
                        cursor.execute('''  
                                   SELECT COUNT(*)  
                                   FROM EquipoEntregadoAdquisicion  
                                   WHERE Estatus = 'EN PROCESO'  
                               ''')
                        count = cursor.fetchone()[0]
                        story.append(Paragraph(f'<b>{question}</b>: {count}', styles['Normal']))
                        paragraph_count += 1

                        # Equipos por contrato
                    elif question == '¿Cuantos equipos de computo hay asignados en cada contrato?':
                        story.append(Paragraph('<b>Equipos por contrato:</b>', styles['Normal']))
                        for Contrato in Contratos:
                            cursor.execute('''  
                                       SELECT COUNT(*)  
                                       FROM ArrendadoRetirar  
                                       WHERE Contrato = ?  
                                   ''', (Contrato,))
                            count = cursor.fetchone()[0]
                            story.append(Paragraph(f'{Contrato}: {count}', styles['Normal']))
                            paragraph_count += 1

                            # Contar tipos de equipos
                    elif question == '¿Cuantos son laptop normal, laptop uso normal, laptop uso rudo, pc uso general, videoproyector fijo, videoproyector portatil y workstation?':
                        cursor.execute('''  
                                   SELECT Componente, COUNT(*)  
                                   FROM EquipoEntregadoAdquisicion  
                                   GROUP BY Componente  
                               ''')
                        resultados = cursor.fetchall()
                        story.append(Paragraph('<b>Componente:</b>', styles['Normal']))
                        for resultado in resultados:
                            story.append(Paragraph(f'{resultado[0]}: {resultado[1]}', styles['Normal']))
                            paragraph_count += 1

                            # Equipos por organismo
                    elif question == '¿Cuantos equipos de computo hay en cada organismo?':
                        cursor.execute('''  
                                   SELECT Organismo, COUNT(*)  
                                   FROM EquipoEntregadoAdquisicion  
                                   GROUP BY Organismo  
                               ''')
                        resultados = cursor.fetchall()
                        story.append(Paragraph('<b>Equipos por organismo:</b>', styles['Normal']))
                        for resultado in resultados:
                            story.append(Paragraph(f'{resultado[0]}: {resultado[1]}', styles['Normal']))
                            paragraph_count += 1

                            # Equipos por área
                    elif question == '¿Cuantos equipos de computo hay en cada area?':
                        cursor.execute('''  
                                   SELECT AreaDepartamento, COUNT(*)  
                                   FROM EquipoEntregadoAdquisicion  
                                   GROUP BY AreaDepartamento  
                               ''')
                        resultados = cursor.fetchall()
                        story.append(Paragraph('<b>Equipos por área:</b>', styles['Normal']))
                        for resultado in resultados:
                            story.append(Paragraph(f'{resultado[0]}: {resultado[1]}', styles['Normal']))
                            paragraph_count += 1

                            # Equipos por subdirección
                    elif question == '¿Cuantos equipos de computo hay en cada subdireccion del activo?':
                        cursor.execute('''  
                                   SELECT Subdireccion, COUNT(*)  
                                   FROM EquipoEntregadoAdquisicion  
                                   GROUP BY Subdireccion  
                               ''')
                        resultados = cursor.fetchall()
                        story.append(Paragraph('<b>Equipos por subdireccion:</b>', styles['Normal']))
                        for resultado in resultados:
                            story.append(Paragraph(f'{resultado[0]}: {resultado[1]}', styles['Normal']))
                            paragraph_count += 1

                            # Controlar el número de párrafos por página
                    if paragraph_count >= MAX_PARAGRAPHS_PER_PAGE:
                        story.append(PageBreak())
                        paragraph_count = 0  # Reinicia el contador para la nueva página

            # Construir el PDF temporal
            doc.build(story)
            conn.close()

            # Combinar el contenido del PDF temporal con el fondo
            self.merge_pdf_with_background(temp_pdf_file, background_pdf, output_pdf_file)

            # Mostrar un mensaje de éxito
            QtWidgets.QMessageBox.information(None, 'Reporte', f'El reporte PDF ha sido generado: {output_pdf_file}')

            # Abrir el PDF automáticamente
            self.open_file(output_pdf_file)

        except Exception as e:
            QtWidgets.QMessageBox.critical(None, 'Error', f'Ocurrió un error: {str(e)}')

    def merge_pdf_with_background(self, report_pdf, background_pdf, output_pdf):
            # Leer el PDF de fondo y el PDF generado
            background = PdfReader(background_pdf)
            report = PdfReader(report_pdf)

            # Verificar que el PDF de fondo tenga al menos una página
            if len(background.pages) == 0:
                raise ValueError("El PDF de fondo no tiene páginas.")

            # Crear un nuevo escritor
            writer = PdfWriter()

            # Combinar cada página del reporte con una copia del fondo
            for page in report.pages:
                # Crear una copia fresca del fondo
                background_page = PageObject.create_blank_page(width=background.pages[0].mediabox.width,
                                                               height=background.pages[0].mediabox.height)
                background_page.merge_page(background.pages[0])  # Añadir el fondo
                background_page.merge_page(page)  # Fusionar con la página del reporte
                writer.add_page(background_page)  # Agregar al escritor

            # Escribir el PDF final
            with open(output_pdf, 'wb') as f:
                writer.write(f)

            # Eliminar el PDF temporal si es necesario
            os.remove(report_pdf)

    def create_excel(self, selected_questions):
        try:
            # Verificar que file_manager esté inicializado
            if not self.file_manager:
                raise ValueError("El atributo 'file_manager' no ha sido inicializado correctamente.")
                # Verificar y limpiar archivos Excel antiguos
            self.file_manager.check_and_clean('excel')

            # Conectar a la base de datos
            conn = sqlite3.connect('equipo_computo.db')
            cursor = conn.cursor()

            # Crear un nuevo archivo Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Reporte"

            # Definir estilos
            title_font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
            subtitle_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            question_font = Font(name='Calibri', size=11, bold=True, color="000000")
            answer_font = Font(name='Calibri', size=11, color="000000")

            # Colores
            header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            subtitle_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            alternate_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            # Bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Alineaciones
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # Título del reporte
            sheet.merge_cells('A1:B2')
            title_cell = sheet['A1']
            title_cell.value = "Reporte de Equipos de Cómputo"
            title_cell.font = title_font
            title_cell.fill = header_fill
            title_cell.alignment = center_align

            # Fecha del reporte
            sheet.merge_cells('A3:B3')
            date_cell = sheet['A3']
            date_cell.value = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            date_cell.font = subtitle_font
            date_cell.fill = subtitle_fill
            date_cell.alignment = center_align

            # Encabezados de columnas
            header_row = 5
            headers = ["Pregunta", "Respuesta"]
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=header_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align

                # Ajustar el ancho de las columnas
            sheet.column_dimensions['A'].width = 55
            sheet.column_dimensions['B'].width = 55

            current_row = header_row + 1
            row_count = 0

            def write_data(question, data):
                nonlocal current_row, row_count
                # Escribir pregunta
                question_cell = sheet.cell(row=current_row, column=1)
                question_cell.value = question
                question_cell.font = question_font
                question_cell.alignment = left_align
                question_cell.border = thin_border

                # Si es una lista de respuestas
                if isinstance(data, list):
                    first = True
                    for item in data:
                        if not first:
                            sheet.cell(row=current_row, column=1).border = thin_border
                        answer_cell = sheet.cell(row=current_row, column=2)
                        answer_cell.value = item
                        answer_cell.font = answer_font
                        answer_cell.alignment = left_align
                        answer_cell.border = thin_border

                        # Aplicar color alternado
                        if row_count % 2:
                            answer_cell.fill = alternate_fill
                            sheet.cell(row=current_row, column=1).fill = alternate_fill

                        current_row += 1
                        row_count += 1
                        first = False
                else:
                    # Si es una respuesta única
                    answer_cell = sheet.cell(row=current_row, column=2)
                    answer_cell.value = data
                    answer_cell.font = answer_font
                    answer_cell.alignment = left_align
                    answer_cell.border = thin_border

                    if row_count % 2:
                        answer_cell.fill = alternate_fill
                        question_cell.fill = alternate_fill

                    current_row += 1
                    row_count += 1

            #Preguntas seleccionadas
            for question in selected_questions:
                if question == '¿Cuantos equipos de computo estan entregados?':
                    cursor.execute('''  
                            SELECT COUNT(*)  
                            FROM EquipoEntregadoAdquisicion  
                            WHERE Estatus = 'ENTREGADO'  
                        ''')
                    count = cursor.fetchone()[0]
                    write_data(question, str(count))

                elif question == '¿Cuantos equipos de computo se han retirado?':
                    cursor.execute('''  
                        SELECT COUNT(*)  
                        FROM RetiroEquipo  
                        WHERE EstatusRetiro IS NOT NULL  
                        AND LENGTH(TRIM(EstatusRetiro)) > 0  
                    ''')
                    count = cursor.fetchone()[0]
                    write_data(question, str(count))

                elif question == '¿Cuantos equipos de computo se han retirado del activo?':
                    cursor.execute('''  
                        SELECT COUNT(*)  
                        FROM RetiroEquipo  
                            WHERE UPPER(TRIM(RetiroFisico)) = 'RETIRADO'  
                    ''')
                    count = cursor.fetchone()[0]
                    write_data(question, str(count))

                elif question == '¿Cuantos equipos de computo se han reasignado?':
                    cursor.execute('''  
                        SELECT COUNT(*)  
                        FROM EquiposReasignados  
                        WHERE UPPER(TRIM(Si_No)) = 'SI'  
                    ''')
                    count = cursor.fetchone()[0]
                    write_data(question, str(count))

                elif question == '¿Cuantos equipos de computo tienen replicador?':
                    cursor.execute('''  
                        SELECT COUNT(DISTINCT SerieReplicador)  
                        FROM EquipoEntregadoAdquisicion  
                        WHERE SerieReplicador IS NOT NULL   
                        AND TRIM(SerieReplicador) != ''  
                    ''')
                    count = cursor.fetchone()[0]
                    write_data(question, str(count))

                elif question == '¿Cuantos equipos de computo estan en proceso?':
                    cursor.execute('''  
                            SELECT COUNT(*)  
                            FROM EquipoEntregadoAdquisicion  
                            WHERE Estatus = 'EN PROCESO'  
                        ''')
                    count = cursor.fetchone()[0]
                    write_data(question, str(count))

                elif question == '¿Cuantos equipos de computo hay asignados en cada contrato?':
                    contratos = ['44', '48', '165', '165/48', '48/165']
                    respuestas = []
                    for contrato in contratos:
                        cursor.execute('''  
                                                SELECT COUNT(*)  
                                                FROM ArrendadoRetirar  
                                                WHERE Contrato = ?  
                                            ''', (contrato,))
                        count = cursor.fetchone()[0]
                        # Asumimos que 'contrato' no tiene espacios adicionales, pero puedes usar .strip() si se requiere
                        respuestas.append(f"Contrato {contrato.strip()}: {count}")
                    write_data(question, respuestas)

                elif question == '¿Cuantos equipos de computo hay en cada organismo?':
                    cursor.execute('''  
                                            SELECT Organismo, COUNT(*)  
                                            FROM EquipoEntregadoAdquisicion  
                                            GROUP BY Organismo  
                                        ''')
                    resultados = cursor.fetchall()
                    # Limpiar y formatear los valores recuperados
                    respuestas = [f"{resultado[0].strip()}: {resultado[1]}" for resultado in resultados]
                    write_data(question, respuestas)

                elif question == '¿Cuantos son laptop normal, laptop uso normal, laptop uso rudo, pc uso general, videoproyector fijo, videoproyector portatil y workstation?':
                    cursor.execute('''  
                                            SELECT Componente, COUNT(*)  
                                            FROM EquipoEntregadoAdquisicion  
                                            GROUP BY Componente   
                                        ''')
                    resultados = cursor.fetchall()
                    # Limpiar y formatear los valores recuperados
                    respuestas = [f"{resultado[0].strip()}: {resultado[1]}" for resultado in resultados]
                    write_data(question, respuestas)


                elif question == '¿Cuantos equipos de computo hay en cada subdireccion del activo?':
                    cursor.execute('''  
                                            SELECT Subdireccion, COUNT(*)  
                                            FROM EquipoEntregadoAdquisicion  
                                            GROUP BY Subdireccion  
                                        ''')
                    resultados = cursor.fetchall()
                    # Limpiar y formatear los valores recuperados
                    respuestas = [f"{resultado[0].strip()}: {resultado[1]}" for resultado in resultados]
                    write_data(question, respuestas)

                elif question == '¿Cuantos equipos de computo hay en cada area?':
                    cursor.execute('''  
                                            SELECT AreaDepartamento, COUNT(*)  
                                            FROM EquipoEntregadoAdquisicion  
                                            GROUP BY AreaDepartamento  
                                        ''')
                    resultados = cursor.fetchall()
                    # Limpiar y formatear los valores recuperados
                    respuestas = [f"{resultado[0].strip()}: {resultado[1]}" for resultado in resultados]
                    write_data(question, respuestas)

            # Agregar pie de página
            footer_row = current_row + 1
            sheet.merge_cells(f'A{footer_row}:B{footer_row}')
            footer_cell = sheet[f'A{footer_row}']
            footer_cell.value = "Documento generado automáticamente por el Sistema de Gestión de Equipos"
            footer_cell.font = Font(italic=True, size=9)
            footer_cell.alignment = center_align

            # Guardar el archivo Excel
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_file = f'reporte_{timestamp}.xlsx'
            workbook.save(excel_file)
            conn.close()

            # Mostrar mensaje de éxito
            QtWidgets.QMessageBox.information(None, 'Éxito',
                                              f'Reporte generado exitosamente:\n{excel_file}')

            # Abrir el archivo Excel
            self.open_file(excel_file)

        except Exception as e:
            QtWidgets.QMessageBox.critical(None, 'Error',
                                           f'Error al generar el reporte:\n{str(e)}')
            print(f"Error: {str(e)}")

    def open_file(self, file_path):
        """Abrir el archivo generado automáticamente."""
        try:
            if os.name == 'nt':  # Windows
                os.startfile(file_path)
            elif os.name == 'posix':  # macOS o Linux
                os.system(f'open "{file_path}"' if sys.platform == 'darwin' else f'xdg-open "{file_path}"')
        except Exception as e:
            QtWidgets.QMessageBox.warning(None, 'Error', f'No se pudo abrir el archivo automáticamente: {e}')

class FileManager:
    def __init__(self):
        self.MAX_PDF_FILES = 2
        self.MAX_EXCEL_FILES = 2
        self.PDF_PATTERN = 'Reporte_*.pdf'  # Define el patrón para buscar archivos PDF
        self.EXCEL_PATTERN = 'reporte_*.xlsx'  # Define el patrón para buscar archivos Excel

    def get_file_list(self, pattern):
        """Obtiene una lista de archivos ordenados por fecha de modificación"""
        files = glob.glob(pattern)  # Busca archivos que coincidan con el patrón
        print(f"Archivos encontrados para el patrón '{pattern}': {files}")  # Depuración
        return sorted(files, key=os.path.getmtime, reverse=True)  # Ordena por fecha más reciente

    def clean_old_files(self, pattern, max_files):
        """Elimina archivos más antiguos si se supera el límite permitido"""
        files = self.get_file_list(pattern)  # Obtener lista de archivos
        if len(files) > max_files:  #Si hay más archivos que los permitidos
            print(f"Se encontraron {len(files)} archivos. Eliminando excedentes...")  # Depuración
            for old_file in files[max_files:]:  # Mantener los más recientes y eliminar el resto
                try:
                    os.remove(old_file)  # Intenta eliminar el archivo antiguo
                    print(f"Archivo eliminado: {old_file}")  # Confirmación de eliminación
                except Exception as e:
                    print(f"Error al eliminar archivo {old_file}: {str(e)}")  # Manejo de errores

    def check_and_clean(self, file_type):
        """Verifica y limpia archivos según el tipo"""
        if file_type == 'pdf':
            self.clean_old_files(self.PDF_PATTERN, self.MAX_PDF_FILES)
        elif file_type == 'excel':
            self.clean_old_files(self.EXCEL_PATTERN, self.MAX_EXCEL_FILES)

class QuestionSelectionDialog(QtWidgets.QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Seleccionar preguntas')
        self.setGeometry(100, 100, 600, 500)
        layout = QtWidgets.QVBoxLayout()

        # Checkbox para seleccionar/deseleccionar todas las preguntas
        self.select_all_checkbox = QtWidgets.QCheckBox("Seleccionar/Deseleccionar todas")
        self.select_all_checkbox.stateChanged.connect(self.toggle_all_checkboxes)
        layout.addWidget(self.select_all_checkbox)

        # Scroll area para las preguntas
        scroll_area = QtWidgets.QScrollArea()
        scroll_widget = QtWidgets.QWidget()
        scroll_layout = QtWidgets.QVBoxLayout()

        # Categorías y preguntas
        self.checkboxes = []  # Lista para almacenar todos los checkboxes
        questions_by_category = {
            "Equipos retirados": [
                "¿Cuantos equipos de computo se han retirado?",
                "¿Cuantos equipos de computo se han retirado del activo?"
            ],
            "Equipos reasignados": [
                "¿Cuantos equipos de computo se han reasignado?"
            ],
            "Estado de los equipos": [
                "¿Cuantos equipos de computo estan entregados?",
                "¿Cuantos equipos de computo estan en proceso?"
            ],
            "Equipos por contrato": [
                "¿Cuantos equipos de computo hay asignados en cada contrato?"
            ],
            "Tipos específicos de equipos": [
                "¿Cuantos son laptop normal, laptop uso normal, laptop uso rudo, pc uso general, videoproyector fijo, videoproyector portatil y workstation?",
                "¿Cuantos equipos de computo tienen replicador?"
            ],
            "Equipos por área o ubicación": [
                "¿Cuantos equipos de computo hay en cada organismo?",
                "¿Cuantos equipos de computo hay en cada area?",
                "¿Cuantos equipos de computo hay en cada subdireccion del activo?"
            ]
        }

        # Crear checkboxes organizados por categoría
        for category, questions in questions_by_category.items():
            # Añadir el título de la categoría
            category_label = QtWidgets.QLabel(f"<b>{category}</b>")
            scroll_layout.addWidget(category_label)

            # Añadir las preguntas de la categoría
            for question in questions:
                checkbox = QtWidgets.QCheckBox(question)
                scroll_layout.addWidget(checkbox)
                self.checkboxes.append(checkbox)

        # Configurar el scroll area
        scroll_widget.setLayout(scroll_layout)
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(scroll_widget)
        layout.addWidget(scroll_area)

        # Botones de aceptar y cancelar
        buttons = QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel
        self.buttonBox = QtWidgets.QDialogButtonBox(buttons)
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.reject)
        layout.addWidget(self.buttonBox)

        self.setLayout(layout)

    def toggle_all_checkboxes(self, state):
        """Seleccionar o deseleccionar todas las preguntas."""
        for checkbox in self.checkboxes:
            checkbox.setChecked(state == QtCore.Qt.Checked)

    def get_selected_questions(self):
        """Obtener las preguntas seleccionadas."""
        return [cb.text() for cb in self.checkboxes if cb.isChecked()]

# Código principal para ejecutar la aplicación
if __name__ == "__main__":

    app = QtWidgets.QApplication(sys.argv)  # Crear la aplicación Qt
    MainWindow = QtWidgets.QMainWindow()  # Crear la ventana principal
    ui = Ui_MainWindow()  # Instanciar la clase de la interfaz
    ui.setupUi(MainWindow)  # Configurar la interfaz
    # Cargar los datos en la tabla
    ui.acciones = Acciones()  # Instanciar Acciones
    ui.acciones.cargar_datos_en_tabla_EquipoEntregadoAdquisicion(ui.tableView)  # Asegúrate de que 'tableView' sea el nombre correcto
    MainWindow.show()  # Mostrar la ventana principal
    sys.exit(app.exec_())  # Ejecutar el bucle de eventos de la aplicación